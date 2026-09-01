from __future__ import annotations

import hashlib
import json
import re
import zipfile
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable


EXPORT_ROOT = Path(__file__).resolve().parent / "data" / "exports"
ALLOWED_SCOPES = {"MASTER", "DEPARTMENT", "PERSON"}


def safe_token(value: str) -> str:
    token = re.sub(r"[^A-Za-z0-9_.-]+", "-", str(value).strip()).strip("-.")
    return token[:80] or "UNASSIGNED"


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def validate_export(operation: dict, release: bool = False) -> dict:
    tasks = operation.get("planning_tasks", [])
    assignments = (operation.get("staffing") or {}).get("assignments", [])
    blockers: list[str] = []
    warnings: list[str] = []
    if not operation.get("planned_start"):
        blockers.append("Operation planned start is not defined.")
    if not tasks:
        blockers.append("No controlled preparation tasks exist.")
    if not assignments:
        blockers.append("No approved team assignments exist.")
    for task in tasks:
        label = task["task_code"]
        if not task.get("assigned_person") or task["assigned_person"].upper() == "UNASSIGNED":
            blockers.append(f"{label}: responsible person is unassigned.")
        if task.get("safety_critical") and task.get("responsible_role") == task.get("verifier_role"):
            blockers.append(f"{label}: safety-critical work lacks independent verification.")
        if task.get("status") == "BLOCKED":
            blockers.append(f"{label}: task is blocked - {task.get('blocker') or 'cause not recorded'}.")
        if task.get("status") != "ACCEPTED":
            warnings.append(f"{label}: status is {task.get('status', 'UNKNOWN').replace('_', ' ')}.")
        if not task.get("evidence"):
            warnings.append(f"{label}: no evidence record is attached.")
    if release:
        blockers.extend(warnings)
    return {
        "release_ready": not blockers,
        "blockers": list(dict.fromkeys(blockers)),
        "warnings": list(dict.fromkeys(warnings)),
        "summary": {
            "tasks": len(tasks),
            "accepted": sum(x.get("status") == "ACCEPTED" for x in tasks),
            "evidence": sum(len(x.get("evidence", [])) for x in tasks),
            "assignments": len(assignments),
        },
    }


def scoped_tasks(operation: dict, scope_kind: str, scope_key: str) -> list[dict]:
    tasks = operation.get("planning_tasks", [])
    if scope_kind == "DEPARTMENT":
        return [x for x in tasks if x["department_code"] == scope_key]
    if scope_kind == "PERSON":
        return [x for x in tasks if x["responsible_role"] == scope_key or x["verifier_role"] == scope_key]
    return list(tasks)


def _paragraph(text: object) -> str:
    return str(text or "-").replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;").replace("\n", "<br/>")


def build_pdf(path: Path, operation: dict, tasks: list[dict], metadata: dict) -> None:
    from reportlab.lib import colors
    from reportlab.lib.enums import TA_CENTER
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    styles = getSampleStyleSheet()
    ink, cyan, line, muted = colors.HexColor("#09202A"), colors.HexColor("#0A9AC1"), colors.HexColor("#B8CBD2"), colors.HexColor("#587681")
    styles.add(ParagraphStyle(name="DocTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=24, leading=29, textColor=ink, spaceAfter=8))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=13, leading=16, textColor=ink, spaceBefore=10, spaceAfter=6))
    styles.add(ParagraphStyle(name="Small", parent=styles["BodyText"], fontSize=7.5, leading=10, textColor=muted))
    styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=7.2, leading=9, textColor=ink))
    styles.add(ParagraphStyle(name="HeaderCell", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=7, leading=9, textColor=colors.white))
    styles.add(ParagraphStyle(name="CoverCode", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=10, leading=13, textColor=cyan, alignment=TA_CENTER))

    def footer(canvas, doc):
        canvas.saveState(); canvas.setStrokeColor(line); canvas.line(16 * mm, 12 * mm, 281 * mm, 12 * mm)
        canvas.setFont("Helvetica", 7); canvas.setFillColor(muted)
        canvas.drawString(16 * mm, 7 * mm, f"CONTROLLED COPY | {metadata['package_code']} REV {metadata['revision']} | {metadata['state']}")
        canvas.drawRightString(281 * mm, 7 * mm, f"Page {doc.page}"); canvas.restoreState()

    doc = SimpleDocTemplate(str(path), pagesize=landscape(A4), rightMargin=16*mm, leftMargin=16*mm, topMargin=14*mm, bottomMargin=18*mm,
                            title=f"{operation['code']} Work Package", author="Stellar Mission & Test Control")
    story = [Spacer(1, 24*mm), Paragraph("STELLAR MISSION & TEST CONTROL", styles["CoverCode"]), Spacer(1, 8*mm),
             Paragraph(_paragraph(operation["title"]), styles["DocTitle"]),
             Paragraph(f"{_paragraph(metadata['scope_label'])} WORK PACKAGE", styles["CoverCode"]), Spacer(1, 10*mm)]
    control = [["DOCUMENT", metadata["package_code"], "REVISION", str(metadata["revision"])],
               ["STATUS", metadata["state"], "GENERATED", metadata["generated_at"]],
               ["OPERATION", operation["code"], "PLANNED START", operation.get("planned_start") or "NOT SET"],
               ["OWNER", operation.get("owner") or "-", "RISK CLASS", operation.get("risk_class") or "-"]]
    table = Table(control, colWidths=[28*mm, 82*mm, 30*mm, 82*mm], rowHeights=10*mm)
    table.setStyle(TableStyle([("GRID",(0,0),(-1,-1),.5,line),("BACKGROUND",(0,0),(0,-1),ink),("BACKGROUND",(2,0),(2,-1),ink),
                               ("TEXTCOLOR",(0,0),(0,-1),colors.white),("TEXTCOLOR",(2,0),(2,-1),colors.white),("FONT",(0,0),(-1,-1),"Helvetica",8),
                               ("FONT",(0,0),(0,-1),"Helvetica-Bold",8),("FONT",(2,0),(2,-1),"Helvetica-Bold",8),("VALIGN",(0,0),(-1,-1),"MIDDLE"),
                               ("LEFTPADDING",(0,0),(-1,-1),7)]))
    story += [table, Spacer(1, 8*mm), Paragraph("This document distributes controlled preparation work. Verify the current revision before use. Evidence and acceptance remain governed by the system record.", styles["Small"]), PageBreak()]

    story += [Paragraph("1. Operation Brief", styles["Section"])]
    brief = [["Mission", operation.get("mission_name")], ["Site", operation.get("site")], ["Objective", operation.get("objective")],
             ["Success criteria", " | ".join(operation.get("success_criteria", []))]]
    brief_table = Table([[Paragraph(_paragraph(a), styles["Cell"]), Paragraph(_paragraph(b), styles["Cell"])] for a,b in brief], colWidths=[38*mm, 205*mm])
    brief_table.setStyle(TableStyle([("GRID",(0,0),(-1,-1),.4,line),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#E8F2F5")),("FONT",(0,0),(0,-1),"Helvetica-Bold",8),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),6)]))
    story += [brief_table, Paragraph("2. Responsibility & Timeline", styles["Section"])]
    rows = [["Task", "Department", "Responsible / Assigned", "Accountable", "Verifier", "Due", "Status"]]
    for task in tasks:
        rows.append([task["task_code"], task["department_code"], f"{task['responsible_role']} / {task['assigned_person']}", task["accountable_role"], task["verifier_role"], task["due_at"], task["status"].replace("_", " ")])
    task_table = Table([[Paragraph(_paragraph(x), styles["HeaderCell"] if row_index == 0 else styles["Cell"]) for x in row] for row_index, row in enumerate(rows)], repeatRows=1,
                       colWidths=[23*mm,23*mm,55*mm,27*mm,24*mm,52*mm,31*mm])
    task_table.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),ink),("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONT",(0,0),(-1,0),"Helvetica-Bold",7),
                                    ("GRID",(0,0),(-1,-1),.35,line),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),4),
                                    ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.white,colors.HexColor("#F3F7F8")])]))
    story += [task_table]
    for index, task in enumerate(tasks, 1):
        story += [PageBreak(), Paragraph(f"{index + 2}. {task['task_code']} - {_paragraph(task['title'])}", styles["Section"])]
        fields = [("Department", task["department_name"]), ("Phase", task["phase"]), ("Priority", task["priority"]),
                  ("Responsible", f"{task['responsible_role']} - {task['assigned_person']}"), ("Accountable", task["accountable_role"]),
                  ("Independent verifier", task["verifier_role"]), ("Planned window", f"{task['planned_start']} to {task['due_at']}"),
                  ("Controlled instruction", task["description"]), ("Required inputs", task["required_inputs"]),
                  ("Acceptance criteria", task["acceptance_criteria"]), ("Required evidence", task["required_evidence"]),
                  ("Current status", task["status"].replace("_", " ")), ("Blocker", task.get("blocker") or "None recorded")]
        detail = Table([[Paragraph(a, styles["Cell"]), Paragraph(_paragraph(b), styles["Cell"])] for a,b in fields], colWidths=[42*mm,201*mm])
        detail.setStyle(TableStyle([("GRID",(0,0),(-1,-1),.35,line),("BACKGROUND",(0,0),(0,-1),colors.HexColor("#E8F2F5")),("FONT",(0,0),(0,-1),"Helvetica-Bold",7.5),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),5)]))
        story += [detail, Paragraph("Evidence and review history", styles["Section"])]
        evidence = [["Evidence", "Type", "Reference", "SHA-256", "Status"]]
        for record in task.get("evidence", []):
            evidence.append([record["title"], record["evidence_type"], record["reference"], record["sha256"][:16], record["status"]])
        if len(evidence) == 1: evidence.append(["No evidence supplied", "-", "-", "-", "OPEN"])
        ev = Table([[Paragraph(_paragraph(x), styles["HeaderCell"] if row_index == 0 else styles["Cell"]) for x in row] for row_index, row in enumerate(evidence)], repeatRows=1, colWidths=[62*mm,32*mm,78*mm,43*mm,28*mm])
        ev.setStyle(TableStyle([("BACKGROUND",(0,0),(-1,0),ink),("TEXTCOLOR",(0,0),(-1,0),colors.white),("GRID",(0,0),(-1,-1),.35,line),("VALIGN",(0,0),(-1,-1),"TOP"),("PADDING",(0,0),(-1,-1),4)]))
        story += [ev]
    doc.build(story, onFirstPage=footer, onLaterPages=footer)


def build_xlsx(path: Path, operation: dict, tasks: list[dict], metadata: dict) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook(); control = wb.active; control.title = "Document Control"
    navy, cyan, pale, white = "09202A", "0A9AC1", "E8F2F5", "FFFFFF"
    control.append(["STELLAR MISSION & TEST CONTROL", ""]); control.merge_cells("A1:B1")
    rows = [("Document", metadata["package_code"]),("Revision",metadata["revision"]),("State",metadata["state"]),("Scope",metadata["scope_label"]),
            ("Operation",operation["code"]),("Title",operation["title"]),("Planned start",operation.get("planned_start") or "NOT SET"),("Generated",metadata["generated_at"])]
    for row in rows: control.append(list(row))
    control["A1"].fill=PatternFill("solid",fgColor=navy); control["A1"].font=Font(color=white,bold=True,size=15); control["A1"].alignment=Alignment(horizontal="center")
    for cell in control["A"][1:]: cell.font=Font(bold=True,color=navy); cell.fill=PatternFill("solid",fgColor=pale)
    control.column_dimensions["A"].width=24; control.column_dimensions["B"].width=80

    task_headers=["Task Code","Title","Department","Responsible Role","Assigned Person","Accountable","Verifier","Phase","Start","Due","Duration h","Priority","Safety Critical","Status","Blocker"]
    ws=wb.create_sheet("Tasks"); ws.append(task_headers)
    for t in tasks: ws.append([t["task_code"],t["title"],t["department_code"],t["responsible_role"],t["assigned_person"],t["accountable_role"],t["verifier_role"],t["phase"],t["planned_start"],t["due_at"],t["duration_hours"],t["priority"],bool(t["safety_critical"]),t["status"],t.get("blocker") or ""])
    instructions=wb.create_sheet("Task Instructions"); instructions.append(["Task Code","Controlled Instruction","Required Inputs","Acceptance Criteria","Required Evidence"])
    evidence=wb.create_sheet("Evidence Register"); evidence.append(["Task Code","Evidence Code","Type","Title","Reference","SHA-256","Supplied By","Supplied At","Status"])
    raci=wb.create_sheet("RACI"); raci.append(["Task Code","Responsible","Accountable","Consulted","Informed","Verifier"])
    for t in tasks:
        instructions.append([t["task_code"],t["description"],t["required_inputs"],t["acceptance_criteria"],t["required_evidence"]])
        raci.append([t["task_code"],t["responsible_role"],t["accountable_role"],", ".join(t.get("consulted_roles",[])),", ".join(t.get("informed_roles",[])),t["verifier_role"]])
        for e in t.get("evidence",[]): evidence.append([t["task_code"],e["evidence_code"],e["evidence_type"],e["title"],e["reference"],e["sha256"],e["supplied_by"],e["supplied_at"],e["status"]])
    timeline=wb.create_sheet("Timeline"); timeline.append(["Milestone","Scheduled At","Owner","Status","Notes"])
    for m in operation.get("planning_milestones",[]): timeline.append([m["name"],m["scheduled_at"],m["owner_role"],m["status"],m["notes"]])
    thin=Side(style="thin",color="B8CBD2")
    for sheet in wb.worksheets:
        sheet.freeze_panes="A2"; sheet.auto_filter.ref=sheet.dimensions
        for cell in sheet[1]: cell.fill=PatternFill("solid",fgColor=navy); cell.font=Font(color=white,bold=True); cell.alignment=Alignment(vertical="center",wrap_text=True)
        for row in sheet.iter_rows():
            for cell in row: cell.border=Border(bottom=thin); cell.alignment=Alignment(vertical="top",wrap_text=True)
        for col in range(1,sheet.max_column+1):
            longest=max((len(str(sheet.cell(r,col).value or "")) for r in range(1,min(sheet.max_row,80)+1)),default=10)
            sheet.column_dimensions[get_column_letter(col)].width=min(max(longest+2,12),48)
        sheet.sheet_view.showGridLines=False
    wb.save(path)


def create_package_files(directory: Path, operation: dict, tasks: list[dict], metadata: dict) -> list[dict]:
    directory.mkdir(parents=True, exist_ok=True)
    stem = safe_token(metadata["package_code"] + "-R" + str(metadata["revision"]))
    pdf_path, xlsx_path, zip_path = directory / f"{stem}.pdf", directory / f"{stem}.xlsx", directory / f"{stem}.zip"
    build_pdf(pdf_path, operation, tasks, metadata); build_xlsx(xlsx_path, operation, tasks, metadata)
    files=[]
    for kind,path,mime in (("PDF",pdf_path,"application/pdf"),("XLSX",xlsx_path,"application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")):
        files.append({"document_type":kind,"filename":path.name,"storage_path":str(path.resolve()),"mime_type":mime,"byte_size":path.stat().st_size,"sha256":sha256_file(path)})
    manifest={"document_control":metadata,"operation":{"code":operation["code"],"title":operation["title"]},"files":[{k:v for k,v in f.items() if k not in {"storage_path"}} for f in files]}
    manifest_path=directory/f"{stem}-manifest.json"; manifest_path.write_text(json.dumps(manifest,indent=2),encoding="utf-8")
    with zipfile.ZipFile(zip_path,"w",zipfile.ZIP_DEFLATED) as archive:
        for included in (pdf_path,xlsx_path,manifest_path): archive.write(included,included.name)
    files.append({"document_type":"ZIP","filename":zip_path.name,"storage_path":str(zip_path.resolve()),"mime_type":"application/zip","byte_size":zip_path.stat().st_size,"sha256":sha256_file(zip_path)})
    return files
