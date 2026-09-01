import tempfile
import unittest
import zipfile
from pathlib import Path

from openpyxl import load_workbook

from stellar_ops.documents import create_package_files, safe_token, scoped_tasks, validate_export


class ControlledDocumentTests(unittest.TestCase):
    def operation(self):
        task = {"task_code":"SAFE-001","title":"Verify exclusion zone","department_code":"SAFE","department_name":"Safety & Range",
                "responsible_role":"RSO","assigned_person":"Omar Al Balushi","accountable_role":"TD","verifier_role":"LCO",
                "task_type":"VERIFICATION","phase":"SITE","planned_start":"2026-09-14T08:00:00Z","due_at":"2026-09-14T12:00:00Z",
                "duration_hours":4,"priority":"CRITICAL","safety_critical":1,"required_inputs":"Site plan","acceptance_criteria":"Zone clear",
                "required_evidence":"Signed checklist","status":"ACCEPTED","blocker":"","description":"Verify the controlled exclusion zone.",
                "evidence":[{"evidence_code":"E-001","evidence_type":"CHECKLIST","title":"Site checklist","reference":"REF-001","sha256":"a"*64,
                             "supplied_by":"Omar Al Balushi","supplied_at":"2026-09-14T12:00:00Z","status":"VERIFIED"}],"consulted_roles":["GND"],"informed_roles":["PROP"]}
        return {"code":"TEST-001","title":"Static Fire Test","mission_name":"Qualification","site":"Test Site","objective":"Validate motor",
                "success_criteria":["Safe test"],"owner":"Test Director","risk_class":"HAZARDOUS","planned_start":"2026-09-15T08:00:00Z",
                "planning_tasks":[task],"staffing":{"assignments":[{"role_code":"RSO"}]},"planning_milestones":[]}

    def test_release_validation_and_scope_filters(self):
        operation = self.operation()
        self.assertTrue(validate_export(operation, release=True)["release_ready"])
        self.assertEqual(len(scoped_tasks(operation,"DEPARTMENT","SAFE")),1)
        operation["planning_tasks"][0]["assigned_person"]="UNASSIGNED"
        self.assertFalse(validate_export(operation, release=True)["release_ready"])

    def test_generated_bundle_contains_auditable_pdf_xlsx_and_manifest(self):
        operation=self.operation(); metadata={"package_code":"TEST-001-MASTER-WP","revision":1,"state":"DRAFT","scope_kind":"MASTER","scope_key":"ALL",
            "scope_label":"MASTER OPERATION","generated_at":"2026-08-19T08:00:00Z","generated_by":"DOCUMENT CONTROL"}
        with tempfile.TemporaryDirectory() as directory:
            files=create_package_files(Path(directory),operation,operation["planning_tasks"],metadata)
            self.assertEqual({x["document_type"] for x in files},{"PDF","XLSX","ZIP"})
            workbook=load_workbook(next(Path(directory).glob("*.xlsx")),read_only=True)
            self.assertEqual(workbook.sheetnames,["Document Control","Tasks","Task Instructions","Evidence Register","RACI","Timeline"])
            with zipfile.ZipFile(next(Path(directory).glob("*.zip"))) as bundle:
                self.assertTrue(any(x.endswith("-manifest.json") for x in bundle.namelist()))
                self.assertTrue(any(x.endswith(".pdf") for x in bundle.namelist()))

    def test_safe_token_removes_path_traversal(self):
        self.assertEqual(safe_token("../../DEPT / SAFE"),"DEPT-SAFE")


if __name__ == "__main__":
    unittest.main()
